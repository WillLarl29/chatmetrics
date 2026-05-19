import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from app.utils.logger import logger

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from app.utils.logger import logger

class ExcelExporter:
    """Clase para exportar DataFrames a archivos Excel profesionales."""
    
    def __init__(self):
        # Definir estilos corporativos
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
        self.header_font = Font(color="FFFFFF", bold=True)
        
        self.title_font = Font(bold=True, size=12)
        self.alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    def export(self, dataframes_dict, output_path):
        """
        Exporta los dataframes con el diseño específico solicitado:
        Hoja 1: Consolidado General (Postgrado y Pregrado en una hoja)
        Hoja 2: Desglosado (Cada subcategoría en una tabla pequeña)
        """
        try:
            logger.info(f"Iniciando exportación a {output_path}")
            wb = openpyxl.Workbook()
            
            # Remover hoja por defecto
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
                
            # 1. Hoja Consolidado General
            ws_general = wb.create_sheet(title="Consolidado General")
            current_row = 1
            
            if 'Postgrado' in dataframes_dict:
                current_row = self._write_consolidado_table(ws_general, dataframes_dict['Postgrado'], "Chattigo Postgrado", current_row)
                current_row += 2 # Espacio en blanco
                
            if 'Pregrado' in dataframes_dict:
                current_row = self._write_consolidado_table(ws_general, dataframes_dict['Pregrado'], "Chattigo Pregrado", current_row)
                current_row += 2
                
            if 'Otros' in dataframes_dict:
                current_row = self._write_consolidado_table(ws_general, dataframes_dict['Otros'], "Chattigo Otros / N/A", current_row)
                current_row += 2
                
            # Leyenda
            self._write_legend(ws_general, 2)
            self._auto_adjust_columns(ws_general)
            
            # 2. Hoja Desglosado
            ws_desglose = wb.create_sheet(title="Desglosado")
            current_row = 1
            
            for key in ['Postgrado', 'Pregrado', 'Otros']:
                if key in dataframes_dict:
                    df = dataframes_dict[key]
                    for idx, row in df.iterrows():
                        # Omitimos la fila de sumas ("Todas las categorias")
                        if "Todas las" in str(row['Categoría']):
                            continue
                            
                        # Escribir cabeceras
                        headers = df.columns.tolist()
                        for col_idx, header in enumerate(headers, 1):
                            cell = ws_desglose.cell(row=current_row, column=col_idx, value=header)
                            cell.font = self.header_font
                            cell.fill = self.header_fill
                            cell.alignment = self.alignment
                            cell.border = self.border
                            
                        current_row += 1
                        
                        # Escribir la fila de datos
                        for col_idx, header in enumerate(headers, 1):
                            val = row[header]
                            cell = ws_desglose.cell(row=current_row, column=col_idx, value=val)
                            cell.alignment = self.alignment
                            cell.border = self.border
                            
                        current_row += 2 # Espacio en blanco entre tablas pequeñas
                        
            self._auto_adjust_columns(ws_desglose)
            
            # Guardar el archivo
            wb.save(output_path)
            logger.info("Exportación completada exitosamente.")
            return True, "Exportación exitosa"
        except Exception as e:
            logger.error(f"Error en exportación: {str(e)}")
            return False, f"Error al exportar: {str(e)}"

    def _write_consolidado_table(self, ws, df, title, start_row):
        """Escribe una tabla consolidada con su título combinado."""
        num_cols = len(df.columns)
        
        # Título
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = self.title_font
        title_cell.alignment = self.alignment
        
        # Poner borde a la celda combinada
        for col_idx in range(1, num_cols + 1):
            ws.cell(row=start_row, column=col_idx).border = self.border
            
        start_row += 1
        
        # Cabeceras
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.title_font
            cell.alignment = self.alignment
            cell.border = self.border
            
        start_row += 1
        
        # Filas de datos
        for idx, row in df.iterrows():
            for col_idx, header in enumerate(headers, 1):
                val = row[header]
                cell = ws.cell(row=start_row, column=col_idx, value=val)
                cell.alignment = self.alignment
                cell.border = self.border
                
                if "Todas las" in str(row['Categoría']):
                    cell.font = Font(bold=True) # Fila de totales en negrita
                    
            start_row += 1
            
        return start_row

    def _write_legend(self, ws, start_row):
        """Escribe la tabla de leyenda al final."""
        legend_data = [
            ("Variable", "Explicación"),
            ("Generación", "Mensaje generado en el sistema"),
            ("Salida", "Mensaje sale de Chattigo a WSP"),
            ("Envío", "Mensaje aceptado por WSP"),
            ("Entrega", "Mensaje llegó a dispositivo"),
            ("Lectura", "Mensaje visualizado"),
            ("Respuesta", "Respuesta Mensaje")
        ]
        
        for i, row_data in enumerate(legend_data):
            for j, val in enumerate(row_data):
                cell = ws.cell(row=start_row + i, column=10 + j, value=val) # Empezar en columna J (10)
                cell.border = self.border
                if i == 0:
                    cell.font = Font(bold=True)
                    cell.alignment = self.alignment
                    
    def _auto_adjust_columns(self, ws):
        """Ajusta automáticamente el ancho de las columnas."""
        from openpyxl.utils import get_column_letter
        
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                # Evitar leer celdas combinadas ya que no tienen length exacto y causan error
                if type(cell).__name__ != 'MergedCell' and cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
                    
            if max_length > 0:
                adjusted_width = (max_length + 2) * 1.2 
                ws.column_dimensions[column_letter].width = adjusted_width
